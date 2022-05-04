# -*- coding: utf-8 -*-
# This file is part of HeCalc, which calculates (U-Th)/He dates and uncertainties
# Copyright (C) 2021 Peter E. Martin <pemartin92@gmail.com>

# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

'''
This module provides funcitonality for either the GUI (through a
threaded worker class using Qthread or through a simple function
to read in data and call the core functionality of HeCalc.

This module is not necessary to run HeCalc if a different data
read-in is desired; the core HeCalc functions may be called indepenently
by other scripts.

If main is used (through the GUI by running the main_GUI.py script
or command line by running this script), it will accept
an excel, csv, or tan-delimited text file. This file must have the headers:
Sample, 238U, 232Th, 147Sm, 4He, 238Ft, 235Ft, 232Ft, and 147Ft.
If 235U was measured directly, this column must also be present.

If run through the command line, each parameter will be set using
the input() built-in function. Files are loaded using tkinter.
The command line version of HeCalc does not support "manual input"
as the GUI does. This is because it is simpler to write a quick
script that calls the individual date calculation and uncertainty
propagation functions contained in this package than it is to include
this option every time HeCalc is run in the command line.

Uncertainty in each value is required (even if the uncertainty is 0).
The uncertainty must follow its respective column; the header for
these uncertainty columns have no labeling requirements.

The output from this module is an excel file with a header containing
the input file path, and (if Monte Carlo uncertainty popagation is used)
the user-requested precision.

The column headers for the output include sample name, time to run, and
corrected and uncorrected dates at a minimum. If linear uncertainty
propagation is requested, 1-s uncertainty columns are added.
If Monte Carlo uncertainty propagation is requested, +/- 68% confidence
intervals are added.
'''

import pandas as pd
import numpy as np
import re
from openpyxl import styles, utils, Workbook
import tkinter
from tkinter.filedialog import askopenfilename, asksaveasfilename
from .date_calculation import get_date
from .linear_propagation import date_uncertainty, date_uncertainty_with235
from .montecarlo import monte_carlo

def _get_cols(linear, monteCarlo, parameterize):
    '''Helper function to create the names of the data to output'''
    save_columns = ['Sample',
                    'Raw date',
                    'Linear raw uncertainty',
                    'MC average CI, raw',
                    'MC +68% CI, raw',
                    'MC -68% CI, raw',
                    'Corrected date',
                    'Linear corrected uncertainty',
                    'MC average CI, corrected',
                    'MC +68% CI, corrected',
                    'MC -68% CI, corrected',
                    'Number of Monte Carlo simulations']
    if monteCarlo and not linear:
        save_columns = ['Sample',
                        'Raw date',
                        'MC average CI, raw',
                        'MC +68% CI, raw',
                        'MC -68% CI, raw',
                        'Corrected date',
                        'MC average, corrected',
                        'MC +68% CI, corrected',
                        'MC -68% CI, corrected',
                        'Number of Monte Carlo simulations']
    elif linear and not monteCarlo:
        save_columns = ['Sample',
                        'Raw date',
                        'Linear raw uncertainty',
                        'Corrected date',
                        'Linear corrected uncertainty']
    elif not monteCarlo and not linear:
        save_columns = ['Sample',
                        'Raw date',
                        'Corrected date']
    # If parameterization is chosen, add the relevant columns
    # at the end of the raw and corrected data output
    if parameterize:
        add_cols = ['Hist raw fit a','Hist raw fit u','Hist raw fit s',
                    'Hist corrected fit a','Hist corrected fit u','Hist corrected fit s']
        save_columns = save_columns+add_cols
    return save_columns

def _load_file(file, measured_U235, sheet):
    '''
    Ensures a user-provided file has the correct columns and loads the data.
    
    The columns can be in any order as long as they are present and have
    the uncertainty in the following column.
    '''
    # List of expected column names
    cols = ['mol 238U',
            'mol 232Th',
            'mol 147Sm',
            'mol 4He',
            '238Ft',
            '235Ft',
            '232Ft',
            '147Ft']
    if measured_U235:
        cols.insert(1, 'mol 235U')
    
    # Check file type and load data
    if file[-5:] == '.xlsx' or file[-4:] == '.xls':
        if not sheet:
            try:
                data_load = pd.read_excel(file)
                # Ensure that all expected columns are present
                assert set(cols).issubset(list(data_load.columns))
            except AssertionError:
                return None
        elif sheet:
            try:
                data_load = pd.read_excel(file, sheet_name = sheet)
                assert set(cols).issubset(list(data_load.columns))
            except:
                return None
    if file[-4:] == '.csv':
        try:
            data_load = pd.read_csv(file)
            assert set(cols).issubset(list(data_load.columns))
        except AssertionError:
            return None
    if file[-4:] == '.txt':
        try:
            data_load = pd.read_csv(file, delimiter="\t")
            assert set(cols).issubset(list(data_load.columns))
        except AssertionError:
            return None
    
    # build list of expected correlation coefficient columns
    corrs = [['r 238U-235U', 'r 235U-238U'],
             ['r 238U-232Th', 'r 232Th-238U'],
             ['r 238U-147Sm', 'r 147Sm-238U'],
             ['r 235U-232Th', 'r 232Th-235U'],
             ['r 235U-147Sm', 'r 147Sm-235U'],
             ['r 232Th-147Sm', 'r 147Sm-232Th'],
             ['r 238Ft-235Ft', 'r 235Ft-238Ft'],
             ['r 238Ft-232Ft', 'r 232Ft-238Ft'],
             ['r 238Ft-147Ft', 'r 147Ft-238Ft'],
             ['r 235Ft-232Ft', 'r 232Ft-235Ft'],
             ['r 235Ft-147Ft', 'r 147Ft-235Ft'],
             ['r 232Ft-147Ft', 'r 147Ft-232Ft']]
    
    # create a list of the column headers and the name of following column
    # to get the uncertainty columns as well
    col_load = ['Sample']
    for c in cols:
        col_load.append(c)
        col_load.append(data_load.columns[data_load.columns.get_loc(c)+1])
    # Add correlation coefficient columns
    for n in corrs:
        for co in n:
            if co in data_load.columns:
                col_load.append(co)
    
    # Create list of column header names to load
    final_cols = ['Sample']
    for i in range(len(cols)):
        final_cols.append(cols[i])
        final_cols.append(u'\u00B1 '+cols[i])
    
    # Restrict data only to relevant data and shorten the
    # column names for easier handling down the road
    data = data_load[col_load]
    data.rename(columns={k:v for k, v in zip(col_load, final_cols)}, inplace = True)
    short_names = {name:name.replace('mol ','') for name in data.columns if 'mol ' in name}
    data = data.rename(columns=short_names)
    
    # Identify which correlation coefficients are present
    cor_cols = []
    missing_cors = []
    for c in corrs:
        for d in c:
            found = False
            if d in data:
                cor_cols.append(d)
                found = True
            if not found:
                missing_cors.append(c[0].split(' ')[-1])
    
    # Convert correlation coefficients to Covariance (if present)
    order_pref = ['238U', '235U','232Th','147Sm', '238Ft', '235Ft', '232Ft', '147Ft']
    for c in cor_cols:
        # Put isotopes and Fts in consistent order
        isos = c.split(' ')[-1].split('-')
        if order_pref.index(isos[0]) > order_pref.index(isos[1]):
            isos = isos[::-1]
        # Convert correlation coefficients to Covariance
        data[isos[0]+'-'+isos[1]] = data[u'\u00B1 '+isos[0]]*data[u'\u00B1 '+isos[1]]*data[c]
        data.drop(c, axis=1, inplace=True)
    # Add missing covariance values to dataframe, default to 0
    data[missing_cors] = 0
    
    data.dropna(axis=0, how = 'all', inplace=True)
    
    return data

def _make_excel(save_out, save_columns, file, monteCarlo, precision_user, saveAs):
    '''
    Helper function to construct the workbook to be saved.
    '''
    # Generate the excel file
    book = Workbook()
    
    # Start by adding histograms to their own sheet and
    # then deleting them to clean up the save_out variable
    if 'raw histogram' in save_out:
        hist_sheet = book.active
        hist_sheet.title = 'Histogram Output'
        keys = {0: 'raw histogram',
                1: 'corrected histogram'}
        xy = {0: 'bin centers',
              1: 'values'}
        for col in range(int(len(save_out['Sample']))):
            for i in range(2):
                for n in range(2):
                    hist_sheet.cell(row=1, column=col*4+i*2+n+1,
                                    value= save_out['Sample'][col]+'\n'+keys[i]+'\n'+xy[n])
                    for row in range(len(save_out[keys[i]][col][n])):
                        hist_sheet.cell(row=row+2, column=col*4+i*2+n+1,
                                        value=save_out[keys[i]][col][n][row])
        del save_out['raw histogram']
        del save_out['corrected histogram']
        # Format histogram column headers
        for header in hist_sheet["1:1"]:
            header.font = styles.Font(bold=True)
            header.alignment = styles.Alignment(wrapText=True)
            if 'raw' in header.value:
                hist_sheet.column_dimensions[utils.get_column_letter(header.column)].width = 13
            if 'corrected' in header.value:
                hist_sheet.column_dimensions[utils.get_column_letter(header.column)].width = 11
       
    # Create a list to format for Excel
    colHead_row = "2:2"
    save_final = []
    # Use the ordering of keys obtained in _get_cols to dictate the order of columns in the output
    # since save_out is a dictionary this is both safer in terms of ordering and allowed columns
    # to be more easily reordered in the future
    for s in save_columns:
        save_final.append(save_out[s])
    save_final = list(zip(*save_final))
    save_final.insert(0, save_columns)
    save_final.insert(0, ['input = ' + file])
    if monteCarlo:
        save_final.insert(0, ['precision = ' + str(round(precision_user,3))+'%'])
        colHead_row = "3:3"
    
    # Make a new sheet for the uncertainty results
    sheet = book.active
    if sheet.title == 'Histogram Output':
        book.create_sheet('Uncertainty Output', 0)
        output = book.active
    else:
        output = book.active
        output.title = 'Uncertainty Output'
    
    # Write the list to the excel file
    for row in range(len(save_final)):
        for col in range(len(save_final[row])):
            output.cell(row=row+1, column=col+1, value=save_final[row][col])
    
    # Format the column headers
    for cell in output[colHead_row]:
        cell.font = styles.Font(bold=True)
        cell.alignment = styles.Alignment(wrapText=True)
        if 'Linear' in cell.value:
            output.column_dimensions[utils.get_column_letter(cell.column)].width = 15
        if 'Corrected date' in cell.value:
            output.column_dimensions[utils.get_column_letter(cell.column)].width = 10
        if 'simulations' in cell.value:
            output.column_dimensions[utils.get_column_letter(cell.column)].width = 17
        if 'average CI' in cell.value:
            output.column_dimensions[utils.get_column_letter(cell.column)].width = 12
        if '% CI, corrected' in cell.value:
            output.column_dimensions[utils.get_column_letter(cell.column)].width = 12
        if 'Hist corrected' in cell.value:
            output.column_dimensions[utils.get_column_letter(cell.column)].width = 13
    output.row_dimensions[3].height = 31
    
    try:
        save_out['histogram']
    except:
        pass
    
    return book

def _linear_propagation(nominal_t, data, measured_U235):
    '''
    Helper function to call the linear uncertainty propagation functions
    including all optional parameters.
    '''
    if measured_U235:
        linear_raw_uncertainty = date_uncertainty_with235(
            data['4He'], nominal_t['raw date'], data[u'\u00B1 4He'],
            data['238U'], data['235U'], data['232Th'], data['147Sm'],
            U238_s=data[u'\u00B1 238U'], U235_s=data[u'\u00B1 235U'],
            Th232_s=data[u'\u00B1 232Th'], Sm147_s=data[u'\u00B1 147Sm'],
            U238_U235_v=data['238U-235U'], U238_Th232_v=data['238U-232Th'],
            U238_Sm147_v=data['238U-147Sm'], U235_Th232_v=data['235U-232Th'],
            U235_Sm147_v=data['235U-147Sm'], Th232_Sm147_v=data['232Th-147Sm']
            )
        linear_corr_uncertainty = date_uncertainty_with235(
            data['4He'], nominal_t['corrected date'], data[u'\u00B1 4He'],
            data['238U'], data['235U'], data['232Th'], data['147Sm'],
            data['238Ft'], data['235Ft'], data['232Ft'], data['147Ft'],
            data[u'\u00B1 238U'], data[u'\u00B1 235U'],
            data[u'\u00B1 232Th'], data[u'\u00B1 147Sm'],
            data[u'\u00B1 238Ft'], data[u'\u00B1 235Ft'],
            data[u'\u00B1 232Ft'], data[u'\u00B1 147Ft'],
            data['238U-235U'], data['238U-232Th'], data['238U-147Sm'],
            data['235U-232Th'], data['235U-147Sm'], data['232Th-147Sm'],
            data['238Ft-235Ft'], data['238Ft-232Ft'], data['238Ft-147Ft'],
            data['235Ft-232Ft'], data['235Ft-147Ft'], data['232Ft-147Ft']
            )
    else:
        linear_raw_uncertainty = date_uncertainty(
            data['4He'], nominal_t['raw date'], data[u'\u00B1 4He'],
            data['238U'], data['238U']/137.818, data['232Th'], data['147Sm'],
            U238_s=data[u'\u00B1 238U'], Th232_s=data[u'\u00B1 232Th'],
            Sm147_s=data[u'\u00B1 147Sm'], U238_Th232_v=data['238U-232Th'],
            U238_Sm147_v=data['238U-147Sm'], Th232_Sm147_v=data['232Th-147Sm']
            )
        linear_corr_uncertainty = date_uncertainty(
            data['4He'], nominal_t['corrected date'], data[u'\u00B1 4He'],
            data['238U'], data['238U']/137.818, data['232Th'], data['147Sm'],
            data['238Ft'], data['235Ft'], data['232Ft'], data['147Ft'],
            data[u'\u00B1 238U'], data[u'\u00B1 232Th'], data[u'\u00B1 147Sm'],
            data[u'\u00B1 238Ft'], data[u'\u00B1 235Ft'],
            data[u'\u00B1 232Ft'], data[u'\u00B1 147Ft'],
            data['238U-232Th'], data['238U-147Sm'], data['232Th-147Sm'],
            data['238Ft-235Ft'], data['238Ft-232Ft'], data['238Ft-147Ft'],
            data['235Ft-232Ft'], data['235Ft-147Ft'], data['232Ft-147Ft']
            )
    return {'raw unc': linear_raw_uncertainty, 'corr unc': linear_corr_uncertainty}

def _sample_loop(save_out, sample_data, measured_U235, linear, monteCarlo,
                 histograms, parameterize, decimals, precision): 
    '''Gets all relevant data for a single samle'''
    # set up reject variable to keep track of whether to run MC
    reject = False
    
    # Create the U235 data
    U235 = sample_data['238U']/137.818
    if measured_U235:
        U235 = sample_data['235U']
    
    # Get the nominal date from the input data
    nominal_t = get_date(sample_data['4He'], sample_data['238U'], U235,
                         sample_data['232Th'], sample_data['147Sm'],
                         sample_data['238Ft'], sample_data['235Ft'],
                         sample_data['232Ft'], sample_data['147Ft'])
    # Reject MC run if nominal date produces NaN or impossibly
    # old dates, as dates >1e10 yrs requires a long time to run
    # and is extremely unlikely to be a legitimate date
    for t in nominal_t.values():
        if t > 1e10:
            reject = True
        elif np.isnan(t):
            reject = True
    save_out['Raw date'].append(round(nominal_t['raw date']/1e6, decimals))
    save_out['Corrected date'].append(round(nominal_t['corrected date']/1e6, decimals))
    
    # Calls helper function _linear_propagation, which unpacks relevant
    # data and calls the correct functions for linear uncertainty propagation
    linear_uncertainty = _linear_propagation(nominal_t, sample_data, measured_U235)
    
    # Reject MC run if linear uncertainty produces NaN
    for u in linear_uncertainty.values():
        if np.isnan(u):
            reject = True
    
    if linear:
        save_out['Linear raw uncertainty'].append(round(linear_uncertainty['raw unc']/1e6,decimals))
        save_out['Linear corrected uncertainty'].append(round(linear_uncertainty['corr unc']/1e6,decimals))
    
    for k in sample_data:
        if u'\u00B1' in k:
            if sample_data[k]<0:
                reject = True
    
    if monteCarlo and not reject:
        # Estimate the number of cycles needed to reach the requested precision
        # TODO implement some kind of safeguard against huge mc_number results for very uncertain data
        s_est = linear_uncertainty['corr unc']
        mean_est = nominal_t['corrected date']
        mc_number = s_est**2/(precision*mean_est)**2
        if mc_number < 5:
            mc_number = 5
        else:
            mc_number = int(mc_number)
        
        # Create U235 data for the monteCarlo function
        if not measured_U235:
            U235 = None
            U235_s = None
        elif measured_U235:
            U235 = sample_data['235U']
            U235_s = sample_data[u'\u00B1 235U']
        
        # Call the monte carlo module
        mc_results = monte_carlo(
            mc_number, sample_data['4He'], sample_data[u'\u00B1 4He'],
            sample_data['238U'], U235, sample_data['232Th'], sample_data['147Sm'],
            sample_data['238Ft'], sample_data['235Ft'], sample_data['232Ft'], sample_data['147Ft'],
            sample_data[u'\u00B1 238U'], U235_s,
            sample_data[u'\u00B1 232Th'], sample_data[u'\u00B1 147Sm'],
            sample_data[u'\u00B1 238Ft'], sample_data[u'\u00B1 235Ft'],
            sample_data[u'\u00B1 232Ft'], sample_data[u'\u00B1 147Ft'],
            sample_data['238U-235U'], sample_data['238U-232Th'], sample_data['238U-147Sm'],
            sample_data['235U-232Th'], sample_data['235U-147Sm'], sample_data['232Th-147Sm'],
            sample_data['238Ft-235Ft'], sample_data['238Ft-232Ft'], sample_data['238Ft-147Ft'],
            sample_data['235Ft-232Ft'], sample_data['235Ft-147Ft'], sample_data['232Ft-147Ft'],
            histogram=histograms, parameterize=parameterize
            )
                
        # Put Monte Carlo results into data framework for saving
        for ft in ['raw', 'corrected']:
            if 1-mc_results[ft+' date']['cycles']/mc_number < precision:
                if ft == 'corrected':
                    save_out['Number of Monte Carlo simulations'].append(mc_number)
                CI_high = (mc_results[ft+' date']['+68% CI']-nominal_t[ft+' date'])
                CI_low = (nominal_t[ft+' date']-mc_results[ft+' date']['-68% CI'])
                mean_CI = np.average([CI_high, CI_low])
                save_out['MC average CI, '+ft].append(round(mean_CI/1e6,decimals))
                save_out['MC +68% CI, '+ft].append(round(CI_high/1e6,decimals))
                save_out['MC -68% CI, '+ft].append(round(CI_low/1e6,decimals))
                if histograms:
                    mc_results[ft+' date']['histogram'][0] = np.around(mc_results[ft+' date']['histogram'][0]/1e6,decimals)
                    mc_results[ft+' date']['histogram'][1] = mc_results[ft+' date']['histogram'][1]
                    save_out[ft+' histogram'].append(mc_results[ft+' date']['histogram'])
                    if parameterize:
                        save_out['Hist '+ft+' fit a'].append(mc_results[ft+' date']['a'])
                        try:
                            save_out['Hist '+ft+' fit u'].append(round(mc_results[ft+' date']['u']/1e6,decimals))
                            save_out['Hist '+ft+' fit s'].append(round(mc_results[ft+' date']['s']/1e6,decimals))
                        except:
                            save_out['Hist '+ft+' fit u'].append(mc_results[ft+' date']['u'])
                            save_out['Hist '+ft+' fit s'].append(mc_results[ft+' date']['s'])
            else:
                reject = True
    if reject:
        for ft in ['raw', 'corrected']:
            if ft == 'corrected':
                save_out['Number of Monte Carlo simulations'].append('NaN')
            save_out['MC average CI, '+ft].append('NaN')
            save_out['MC +68% CI, '+ft].append('NaN')
            save_out['MC -68% CI, '+ft].append('NaN')
            if histograms:
                save_out[ft+' histogram'].append([np.array(['NaN']),np.array(['NaN'])])
                if parameterize:
                    save_out['Hist, '+ft+' fit a'].append('NaN')
                    save_out['Hist, '+ft+' fit u'].append('NaN')
                    save_out['Hist, '+ft+' fit s'].append('NaN')
    return save_out

def hecalc_main(file=None, saveAs=None, percent_precision=0.01, decimals=2, measured_U235=False,
                monteCarlo=True, linear=True, histograms=False, parameterize=False):
    '''
    Read in a file with He, radionuclide, and Ft data and their uncertainties
    and perform data reduction and uncertainty propagation.
    
    Parameters
    ----------
    file : None, path to file, optional
        If a path is provided, HeCalc will load the given path. Input
        files must be .xlsx, .xls, .csv, or tab-delimited .txt.
        The input file for this module must have the following headers:
        Sample, mol 238U, mol 232Th, mol 147Sm, mol 4He,
        238Ft, 235Ft, 232Ft, and 147Ft
        Sample is the sample name
        mol 238U, mol 232Th, mol 147Sm, and mol 4He are amounts of each
        nuclide; these can be in any self-consistent unit (e.g., mol, fmol, atoms, mol/g)
        238Ft, 235Ft, 232Ft, and 147Ft are the isotope-specific
        Ft values for each nuclide. see Ketcham et al., 2011 for
        how these may be calculated.
        If file is None (the default), a Tkinter popup will prompt
        an input for data reduction
        
    saveAs : None, path, optional
        If a path is provided, HeCalc will save the results
        to this location.
        If saveAs is None (the default), a Tkinter popup will prompt
        an input for save destination
        
    percent_precision: float, optional
        the precision of the Monte Carlo mean determines the number
        of Monte Carlo simulations to perform; the number input here
        is the relative precision in percent (i.e., for a sample
        with a mean age of 50 Ma, a 1% precision would be +/- 0.5 Ma)
        
    decimals: int, optional
        the number of decimals to report. This simply dictates the
        level to which the code will round the result; this
        has no bearing on the statistical portions of the code
        
    measured_U235: bool, optional
        Whether 235U was measured directly. This should only
        be selected in the very rare case that 235U is not
        assumed from 238U measurement
        
    monteCarlo: bool, optional
        Whether to run Monte Carlo uncertainty propagation
        
    linear: bool, optional
        Whether to run linear uncertainty propagation
        
    histograms: bool, optional
        Whether to generate histograms if a Monte Carlo model is run.
    
    parameterize: bool, optional
        Whether to parameterize the histograms with a Skew-normal function.
        
    Returns
    -------
    Saves out an excel file at the path sepcified by saveAs. Results
    for both raw and alpha ejection corrected calculations are included.
    A header giving the requested precision (if Monte Carlo modeling
    was chosen) and the path for the source file
    For all calculations, the sample name and date are returned
    If linear is chosen, the 1-sigma uncertainty is included
    If montecarlo is chosen, the 68% confidence intervals, mean date,
    and number of cycles necessary to reach the requested precision.
    If parameterization was requested, the fitted parameters a, u, and s
    are included with a being shape, u the location parameter, and
    s the scale parameter for a Skew-normal function.
    If histograms are produced, these will be saved in a second excel
    sheet with the both the raw and corrected histograms given with
    bin centers (in units of Ma) and the absolute number of Monte Carlo
    simulations corresponding to each date.
    '''

    # Convert user-defined precision to proportion
    precision = percent_precision/100
    
    # Create the list of columns to be used as headers in the output
    # these vary based on which (if any) uncertainty propagation is
    # chosen by the user.
    save_columns = _get_cols(linear, monteCarlo, parameterize)
    save_out = {c: [] for c in save_columns}
    
    # Generate additional variables to save the histograms
    if histograms:
        save_out['raw histogram'] = []
        save_out['corrected histogram'] = []
    
    # Get the file containing the data if no file was provided
    if file == None:
        root = tkinter.Tk()
        file = askopenfilename(title = 'Choose file to reduce',
                               filetypes=[("Excel files", "*.xlsx .*xls"),
                                          ("Comma-separated values", "*.csv"),
                                          ("Tab-delimited text", "*.txt")])
        root.wm_withdraw()
    
    # Get the path to save to if none provided
    if saveAs == None:
        root = tkinter.Tk()
        saveAs = asksaveasfilename(filetypes=[("Excel files", "*.xlsx .*xls")],
                                     defaultextension=".xlsx")
        root.wm_withdraw()
    
    # Load in the data. The helper function _load_file goes through
    # a series of checks to make sure the file has the appropriate
    # information. If not, an input is requested in case an excel
    # workbook with multiple sheets has been provided
    data = _load_file(file, measured_U235, None)
    while data is None:
        sheet = input('Name of sheet to read data from: ')
        data = _load_file(file, measured_U235, sheet)
        if data is None:
            print('Ensure columns are correctly labeled')
    
    # This is the main loop that calculates progress and calls the main
    # _sample_loop function, which takes the relevant inputs and calls
    # the individual HeCalc functions needed
    for i in range(len(data)):
        # Update user of new sample being analyzed
        print(('Now working on sample ' + str(data['Sample'].iloc[i])))
        save_out['Sample'].append(data['Sample'].iloc[i])

        # Check that no illegal characters are being used and replace with underscores
        if re.sub('[^\w\-_]', '_', data['Sample'].iloc[i]) != data['Sample'].iloc[i]:
            data['Sample'].iloc[i] = re.sub('[^\w\-_]', '_', data['Sample'].iloc[i])
        
        # Get the individual sample data as a dictionary
        sample_data = data.loc[[i]].to_dict(orient='records')[0]
        
        save_out = _sample_loop(save_out, sample_data, measured_U235, linear, monteCarlo,
                                histograms, parameterize, decimals, precision)

    # Create the excel workbook using the input parameters
    book = _make_excel(save_out, save_columns, file, monteCarlo, percent_precision, saveAs)
    
    # Save the excel workbook
    book.save(saveAs)